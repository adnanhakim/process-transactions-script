"""
utils.files
~~~~~~~~~~~~~~

This module contains common file methods.

"""

import logging

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._write_only import WriteOnlyWorksheet


def read_excel_to_list(file_name):
    try:
        # Load the workbook
        workbook: Workbook = load_workbook(file_name)

        # Select the active worksheet
        sheet: ReadOnlyWorksheet | None = workbook.active

        # Read the data into a list of lists
        data: list = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        return data
    except FileNotFoundError:
        logging.error("No such file exists: %s", file_name)
        return []
    except Exception as e:
        logging.error("An error occurred: %s", e)
        return []


def write_list_to_excel(data: list[list[str]], workbook_name: str, sheet_name: str):
    try:
        # Create a new workbook
        workbook: Workbook = Workbook()

        # Select the active worksheet
        sheet: WriteOnlyWorksheet | None = workbook.active
        sheet.title = sheet_name

        # Write the data
        for row in data:
            sheet.append(row)

        workbook.save(workbook_name)
    except Exception as e:
        logging.error("An error occurred: %s", e)
        return []
